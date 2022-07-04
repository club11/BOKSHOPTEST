from unicodedata import name
from django.shortcuts import render
from django.views.generic import FormView
from . import forms
from books import models as books_models
from directories import models as directories_models
from django.http import HttpResponseRedirect
from django.urls import reverse_lazy
from django.contrib.auth.mixins import PermissionRequiredMixin
"""
для работы с excel
"""
# os
import os
# для образений в БД
import sqlite3
# для работы с xls
import openpyxl

class ExcelStaffView(PermissionRequiredMixin, FormView):
    template_name = 'staff/excel_import.html'
    form_class = forms.ExcelStaffForm
    #success_url = reverse_lazy('books:home')
    permission_required = 'directories.add_serie'

    def get_initial(self):
        print(super().get_initial())
        return super().get_initial()

    def form_valid(self, form):
        if self.request.method == 'POST':
            form = forms.ExcelStaffForm(self.request.POST, self.request.FILES)
            if form.is_valid():
                file_to_read = openpyxl.load_workbook(self.request.FILES['file_fields'], data_only=True)     
                sheet = file_to_read['Sheet1']      # Читаем файл и лист1 книги excel
                for row in range(2, sheet.max_row + 1): # Цикл по строкам начиная со второй (в первой заголовки)
                    data = []
                    for col in range(1, 16): # Цикл по столбцам от 1 до 15 ( 16 не включая)
                        value = sheet.cell(row, col).value # value содержит значение ячейки с координатами row col
                        data.append(value)
                    print(data, len(data))
                    author = directories_models.Author.objects.update_or_create(author=data[3])
                    serie = directories_models.Serie.objects.update_or_create(serie=data[4])
                    genre = directories_models.Genre.objects.update_or_create(genre=data[5])
                    editor = directories_models.Editor.objects.update_or_create(editor=data[6])
                    books_models.Book.objects.update_or_create(book_name=data[0], 
                    price=data[1], 
                    currency_price=data[2], 
                    author=directories_models.Author.objects.get(author=data[3]), 
                    serie=directories_models.Serie.objects.get(serie=data[4]), 
                    genre=directories_models.Genre.objects.get(genre=data[5]), 
                    editor=directories_models.Editor.objects.get(editor=data[6]), 
                    publishing_date=data[7], 
                    pages=data[8], 
                    binding=data[9], 
                    format=data[10], 
                    isbn=data[11], 
                    weigh=data[12], 
                    age_restrictions=data[13], 
                    value_available=data[14],
                    )
                return HttpResponseRedirect(reverse_lazy('books:book_list'))
        else:
            form = forms.ExcelStaffForm()
        return render(self.request, 'upload.html', {'form': form})






